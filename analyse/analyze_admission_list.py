#!/usr/bin/env python3
"""Sort admission list by undergraduate school tier and color-code serial numbers."""

import csv
import re
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

INPUT_PATH = Path(
    r"d:\xwechat_files\wxid_qv4v4ol18wpw22_f52e\msg\file\2026-06\第一轮公示名单.xlsx"
)
OUT_DIR = Path(__file__).parent
LIST_DIR = OUT_DIR

# tier: 1=985, 2=211, 3=双一流(双非), 4=四非
TIER_LABEL = {1: "985", 2: "211", 3: "双一流", 4: "四非"}
TIER_FILL = {
    1: PatternFill("solid", fgColor="FFC7CE"),  # light red
    2: PatternFill("solid", fgColor="BDD7EE"),  # light blue
    3: PatternFill("solid", fgColor="D9D9D9"),  # silver gray
    4: PatternFill("solid", fgColor="F2F2F2"),  # lighter gray for 四非
}

# Aliases for renamed / campus schools
ALIASES = {
    "第二军医大学": "海军军医大学",
    "第四军医大学": "空军军医大学",
}


def load_school_set(csv_name: str) -> set[str]:
    schools = set()
    with (LIST_DIR / csv_name).open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            schools.add(normalize(row["学校"]))
    return schools


def normalize(name: str) -> str:
    if not name:
        return ""
    name = str(name).strip()
    for src, dst in [("(", "（"), (")", "）"), (" ", ""), ("　", "")]:
        name = name.replace(src, dst)
    return name


def resolve_tier(school_raw: str, uni985: set, uni211: set, dfc: set) -> int:
    school = normalize(school_raw)
    school = normalize(ALIASES.get(school_raw, school_raw) if school_raw else school)

    # Direct match
    if school in uni985:
        return 1
    if school in uni211:
        return 2
    if school in dfc:
        return 3

    # Campus / branch: e.g. 中国石油大学（北京）克拉玛依校区
    for base in sorted(uni211 | uni985 | dfc, key=len, reverse=True):
        if school.startswith(base) and len(school) > len(base):
            if base in uni985:
                return 1
            if base in uni211:
                return 2
            if base in dfc:
                return 3

    return 4


def read_rows(path: Path) -> tuple[list[str], list[dict]]:
    wb = load_workbook(path)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not vals[0]:
            continue
        row = dict(zip(headers, vals))
        rows.append(row)
    return headers, rows


def sort_key(row: dict, uni985, uni211, dfc) -> tuple:
    tier = resolve_tier(row.get("本科学校", ""), uni985, uni211, dfc)
    return (row.get("专业代码", ""), row.get("专业名称", ""), tier, row.get("报名号", ""))


def write_output(headers: list, rows: list[dict], uni985, uni211, dfc, out_path: Path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "排序结果"

    out_headers = ["序号", "院校层次"] + headers
    ws.append(out_headers)

    header_font = Font(bold=True)
    for c in range(1, len(out_headers) + 1):
        cell = ws.cell(1, c)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    current_major = None
    seq = 0
    stats = {1: 0, 2: 0, 3: 0, 4: 0}
    tier_by_school: dict[str, int] = {}

    for row in rows:
        major_key = (row.get("专业代码"), row.get("专业名称"))
        if major_key != current_major:
            current_major = major_key
            seq = 0
        seq += 1

        school = row.get("本科学校", "")
        tier = resolve_tier(school, uni985, uni211, dfc)
        stats[tier] += 1
        tier_by_school.setdefault(school, tier)

        values = [seq, TIER_LABEL[tier]] + [row.get(h, "") for h in headers]
        ws.append(values)
        row_idx = ws.max_row

        num_cell = ws.cell(row_idx, 1)
        num_cell.fill = TIER_FILL[tier]
        num_cell.alignment = Alignment(horizontal="center")
        num_cell.font = Font(bold=True)

        tier_cell = ws.cell(row_idx, 2)
        tier_cell.alignment = Alignment(horizontal="center")

    # column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 10
    for col, w in zip("CDEFGH", [12, 18, 14, 10, 6, 28]):
        ws.column_dimensions[col].width = w

    # stats sheet
    ws2 = wb.create_sheet("统计")
    ws2.append(["院校层次", "人数", "说明"])
    ws2.append(["985", stats[1], "序号浅红"])
    ws2.append(["211", stats[2], "序号浅蓝"])
    ws2.append(["双一流", stats[3], "双非（非985/211的一流）序号银灰"])
    ws2.append(["四非", stats[4], "非985/211/双一流，放最后"])
    ws2.append([])
    ws2.append(["专业代码", "专业名称", "985", "211", "双一流", "四非", "合计"])
    from collections import defaultdict

    major_stats = defaultdict(lambda: {1: 0, 2: 0, 3: 0, 4: 0})
    for row in rows:
        key = (row.get("专业代码"), row.get("专业名称"))
        t = resolve_tier(row.get("本科学校", ""), uni985, uni211, dfc)
        major_stats[key][t] += 1
    for (code, name), ms in sorted(major_stats.items()):
        total = sum(ms.values())
        ws2.append([code, name, ms[1], ms[2], ms[3], ms[4], total])

    ws3 = wb.create_sheet("学校分层")
    ws3.append(["本科学校", "院校层次"])
    for school in sorted(tier_by_school, key=lambda s: (tier_by_school[s], s)):
        ws3.append([school, TIER_LABEL[tier_by_school[school]]])

    wb.save(out_path)
    return stats, tier_by_school


def main():
    uni985 = load_school_set("985_universities.csv")
    uni211 = load_school_set("211_universities.csv")
    dfc = load_school_set("double_first_class.csv")
    # 211 includes all 985; 双一流 includes most 211
    uni211_only = uni211 - uni985
    dfc_only = dfc - uni211

    headers, rows = read_rows(INPUT_PATH)
    rows.sort(key=lambda r: sort_key(r, uni985, uni211, dfc))

    out_path = OUT_DIR / "第一轮公示名单_分析.xlsx"
    stats, tier_by_school = write_output(
        headers, rows, uni985, uni211, dfc, out_path
    )

    print(f"输入: {INPUT_PATH}")
    print(f"输出: {out_path}")
    print(f"共 {len(rows)} 人")
    for t in (1, 2, 3, 4):
        print(f"  {TIER_LABEL[t]}: {stats[t]}")
    unknown = [s for s, t in tier_by_school.items() if t == 4]
    print(f"四非学校 ({len(unknown)}): {', '.join(unknown)}")


if __name__ == "__main__":
    main()
