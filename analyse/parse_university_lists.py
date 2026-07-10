#!/usr/bin/env python3
"""Parse 985 / 211 / 双一流 lists from dxsbb.com article."""

import csv
import json
import re
from html import unescape
from pathlib import Path

try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None

SOURCE_URL = "https://www.dxsbb.com/news/143601.html"
HTML_PATH = Path(__file__).parent / "_source.html"
OUT_DIR = Path(__file__).parent


def cell_text(html: str) -> str:
    text = re.sub(r"<[^>]+>", "", html)
    return unescape(text).strip()


def parse_tables(html: str) -> list[list[list[str]]]:
    content_m = re.search(r'<div class="content">(.*?)</div>\s*<div class="mark">', html, re.S)
    if not content_m:
        raise ValueError("content div not found")
    content = content_m.group(1)
    tables = []
    for table_html in re.findall(r"<table[^>]*>(.*?)</table>", content, re.S):
        rows = []
        for tr in re.findall(r"<tr>(.*?)</tr>", table_html, re.S):
            cells = [cell_text(td) for td in re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", tr, re.S)]
            if cells:
                rows.append(cells)
        if rows:
            tables.append(rows)
    return tables


def group_by_province(items: list[dict]) -> dict[str, list[str]]:
    grouped: dict[str, list[str]] = {}
    for item in items:
        prov = item["province"]
        grouped.setdefault(prov, []).append(item["name"])
    return grouped


def export_table(path_base: Path, rows: list[dict], name_key: str = "name") -> None:
    """Export 序号/学校 two-column CSV and XLSX."""
    csv_path = path_base.with_suffix(".csv")
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["序号", "学校"])
        for row in rows:
            writer.writerow([row["no"], row[name_key]])

    if Workbook is None:
        return
    xlsx_path = path_base.with_suffix(".xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = path_base.stem
    ws.append(["序号", "学校"])
    for row in rows:
        ws.append([row["no"], row[name_key]])
    wb.save(xlsx_path)


def main():
    html = HTML_PATH.read_text(encoding="utf-8")
    tables = parse_tables(html)
    if len(tables) < 3:
        raise ValueError(f"expected 3 tables, got {len(tables)}")

    # --- 985 ---
    uni985 = []
    last_province = ""
    for row in tables[0][1:]:
        if len(row) < 3:
            continue
        no, name, province = row[0], row[1], row[2]
        if province:
            last_province = province
        else:
            province = last_province
        uni985.append({"no": int(no), "name": name, "province": province})

    # --- 211 ---
    uni211 = []
    for row in tables[1][1:]:
        if len(row) < 3:
            continue
        no, name, province = row[0], row[1], row[2]
        uni211.append({"no": int(no), "name": name, "province": province})

    # --- 双一流 ---
    double_first = []
    for row in tables[2][1:]:
        if len(row) < 3:
            continue
        no, name, disciplines = row[0], row[1], row[2]
        double_first.append({"no": int(no), "name": name, "disciplines": disciplines})

    data = {
        "source": SOURCE_URL,
        "summary": {
            "985_count": len(uni985),
            "211_count": len(uni211),
            "double_first_class_count": len(double_first),
        },
        "985": uni985,
        "211": uni211,
        "double_first_class": double_first,
    }
    (OUT_DIR / "universities.json").write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    # --- Markdown: 985 ---
    lines985 = [
        "# 985 工程大学名单",
        "",
        f"数据来源: [{SOURCE_URL}]({SOURCE_URL})",
        "",
        f"共 **{len(uni985)}** 所。",
        "",
        "## 按序号",
        "",
        "| 序号 | 学校 | 所在地 |",
        "| --- | --- | --- |",
    ]
    for u in uni985:
        lines985.append(f"| {u['no']} | {u['name']} | {u['province']} |")
    lines985.extend(["", "## 按省份", ""])
    for prov, names in sorted(group_by_province(uni985).items(), key=lambda x: -len(x[1])):
        lines985.append(f"### {prov}（{len(names)}所）")
        lines985.append("")
        for n in names:
            lines985.append(f"- {n}")
        lines985.append("")
    (OUT_DIR / "985_universities.md").write_text("\n".join(lines985), encoding="utf-8")

    # --- Markdown: 211 ---
    lines211 = [
        "# 211 工程大学名单",
        "",
        f"数据来源: [{SOURCE_URL}]({SOURCE_URL})",
        "",
        f"共 **{len(uni211)}** 所（含全部 985 高校）。",
        "",
        "## 按序号",
        "",
        "| 序号 | 学校 | 所在地 |",
        "| --- | --- | --- |",
    ]
    for u in uni211:
        lines211.append(f"| {u['no']} | {u['name']} | {u['province']} |")
    lines211.extend(["", "## 按省份", ""])
    for prov, names in sorted(group_by_province(uni211).items(), key=lambda x: -len(x[1])):
        lines211.append(f"### {prov}（{len(names)}所）")
        lines211.append("")
        for n in names:
            lines211.append(f"- {n}")
        lines211.append("")
    (OUT_DIR / "211_universities.md").write_text("\n".join(lines211), encoding="utf-8")

    # --- Markdown: 双一流 ---
    linesdfc = [
        "# 双一流建设高校名单",
        "",
        f"数据来源: [{SOURCE_URL}]({SOURCE_URL})",
        "",
        f"共 **{len(double_first)}** 所。",
        "",
        "| 序号 | 学校 | 双一流建设学科 |",
        "| --- | --- | --- |",
    ]
    for u in double_first:
        disc = u["disciplines"].replace("|", "\\|")
        linesdfc.append(f"| {u['no']} | {u['name']} | {disc} |")
    (OUT_DIR / "double_first_class.md").write_text("\n".join(linesdfc), encoding="utf-8")

    export_table(OUT_DIR / "985_universities", uni985)
    export_table(OUT_DIR / "211_universities", uni211)
    export_table(OUT_DIR / "double_first_class", double_first)

    # combined xlsx with 3 sheets
    if Workbook is not None:
        wb = Workbook()
        wb.remove(wb.active)
        for sheet_name, rows in [
            ("985", uni985),
            ("211", uni211),
            ("双一流", double_first),
        ]:
            ws = wb.create_sheet(sheet_name)
            ws.append(["序号", "学校"])
            for row in rows:
                ws.append([row["no"], row["name"]])
        wb.save(OUT_DIR / "universities.xlsx")

    # --- Combined summary ---
    only211 = [u["name"] for u in uni211 if u["name"] not in {x["name"] for x in uni985}]
    only_dfc = [
        u["name"]
        for u in double_first
        if u["name"] not in {x["name"] for x in uni211}
    ]
    summary = [
        "# 985 / 211 / 双一流 名单汇总",
        "",
        f"数据来源: [{SOURCE_URL}]({SOURCE_URL})",
        "",
        "## 数量概览",
        "",
        "| 类别 | 数量 | 说明 |",
        "| --- | --- | --- |",
        f"| 985 工程 | {len(uni985)} 所 | 重点建设世界一流大学 |",
        f"| 211 工程 | {len(uni211)} 所 | 全部 985 均入选 211 |",
        f"| 双一流 | {len(double_first)} 所 | 含全部 985、211，另增补 {len(only_dfc)} 所（含军改更名等） |",
        "",
        "## 分文件",
        "",
        "- [985_universities.md](./985_universities.md)",
        "- [211_universities.md](./211_universities.md)",
        "- [double_first_class.md](./double_first_class.md)",
        "- [universities.json](./universities.json)",
        "- [985_universities.csv](./985_universities.csv) / [985_universities.xlsx](./985_universities.xlsx)",
        "- [211_universities.csv](./211_universities.csv) / [211_universities.xlsx](./211_universities.xlsx)",
        "- [double_first_class.csv](./double_first_class.csv) / [double_first_class.xlsx](./double_first_class.xlsx)",
        "- [universities.xlsx](./universities.xlsx)（三表合一）",
        "",
        f"## 仅 211、非 985（{len(only211)} 所）",
        "",
    ]
    for n in only211:
        summary.append(f"- {n}")
    summary.extend(["", f"## 仅双一流、非 211（{len(only_dfc)} 所）", ""])
    for n in only_dfc:
        summary.append(f"- {n}")
    summary.append("")
    (OUT_DIR / "README.md").write_text("\n".join(summary), encoding="utf-8")

    print(f"985: {len(uni985)}, 211: {len(uni211)}, 双一流: {len(double_first)}")
    print(f"仅211非985: {len(only211)}, 仅双一流非211: {len(only_dfc)}")


if __name__ == "__main__":
    main()
